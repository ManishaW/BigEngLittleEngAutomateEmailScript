import openpyxl
import textwrap

wb = openpyxl.load_workbook('Match Sheet.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet1')

cellA = "36" #change (from match sheet)
SheMentor= False #change (from intuition lol)
SheMentee =False
cellB= "25" #Big eng Mentor num change
cellC="12" #little eng mentor num change


mentee = sheet['A'+cellA].value.replace(" *","").replace("*","")
mentor = sheet['D'+cellA].value.replace(" *","").replace("*","")
mentor_email = sheet['E'+cellA].value
mentee_email = sheet['B'+cellA].value

#big eng sheet
wb = openpyxl.load_workbook('bigeng.xlsx')
wb.get_sheet_names()
big_sheet = wb.get_sheet_by_name('Form Responses 1')

#little eng sheet
wb = openpyxl.load_workbook('littleeng.xlsx')
wb.get_sheet_names()
little_sheet = wb.get_sheet_by_name('Form Responses 1')


mentor_year= big_sheet['E'+cellC].value
mentee_year= little_sheet['E'+cellC].value.lower()
mentor_prog=big_sheet['D'+cellB].value
mentee_prog=little_sheet['D'+cellC].value
mentor_hobbies=big_sheet['F'+cellB].value.lower()
mentee_hobbies=little_sheet['F'+cellC].value.lower()
mentor_transform =big_sheet['G'+cellB].value.lower()
mentee_transform = little_sheet['G'+cellC].value.lower()
mentor_bucket =big_sheet['H'+cellB].value.lower()
mentee_bucket = little_sheet['H'+cellC].value.lower()
mentor_music =big_sheet['I'+cellB].value.lower()
mentee_music = little_sheet['I'+cellC].value.lower()

if SheMentor is True:
    pronoun = "She"
    pronoun2 = "her"
else:
    pronoun = "He"
    pronoun2 = "his"

if SheMentee is True:
    mentee_pronoun = "She"
    mentee_pronoun2 = "her"
else:
    mentee_pronoun = "He"
    mentee_pronoun2 = "his"
    
print mentor_email
print mentee_email

print textwrap.dedent("""\
Hi %s and %s,

Congratulations! You two have been matched together for our Big Eng Little Eng Mentorship program! I hope you got to meet each other during the kick-off but if not, it's never too late! Try to keep in touch starting today. We hope to see you both at our next event! (Keep your eyes open for an email from me later this month)

-About your mentor-
%s: %s
%s is a %s %s student. %s loves %s! %s believes that the advancements in %sis something that will transform our future. An interesting thing on %s bucket list is that %s wants to %s. How exciting! %s also loves %s music.

      """) % (mentee.split()[0],mentor.split()[0], mentor, mentor_email, mentor.split()[0], mentor_year, mentor_prog,pronoun,mentor_hobbies,pronoun,mentor_transform,pronoun2, pronoun.lower(), mentor_bucket, pronoun,mentor_music)

print textwrap.dedent("""     
-About your mentee-
%s: %s
%s is a %s %s student. %s loves %s! %s believes that the advancements in %s is something that will transform our future. An interesting thing on %s bucket list is that %s wants to %s. Sounds like fun! %s also loves %s music.
Feel free to email me if you have any questions. See you two soon!

-----------------------------------
On another note,

Are you interested in Power Electronics? IEEE Ottawa Section is hosting a Solantro Lab Tour for you! See the flyer attached. It's happening on October 10th!
      
      """) % (mentee, mentee_email, mentee.split()[0], mentee_year, mentee_prog, mentee_pronoun, mentee_hobbies, mentee_pronoun, mentee_transform, mentee_pronoun2, mentee_pronoun, mentee_bucket, mentee_pronoun, mentee_music)
